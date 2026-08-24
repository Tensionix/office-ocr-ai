from __future__ import annotations

import copy
import hashlib
import json
import sys
import zipfile
from pathlib import Path

import fitz
from PIL import Image, ImageDraw
from docx import Document
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from system_core.document_exporters import _balanced_column_widths, export_document_model
from system_core.document_model import create_document_model, load_document_model
from system_core.ocr_brick.engines.regions import Cell, TableRegion
from system_core.ocr_brick.ocr_fusion import fuse_mistral_yandex, needs_yandex_review
from system_core.ocr_brick.table_geometry import recover_physical_tables
from system_core.ocr_brick.engines.base import OcrResult, Word
from system_core.ocr_brick.pipeline_controller import PipelineController


def _sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_lossless_package_and_all_exporters(tmp_path: Path) -> None:
    source = tmp_path / "исходный скан.png"
    image = Image.new("RGB", (1200, 800), "white")
    ImageDraw.Draw(image).text((100, 100), "Audit 123", fill="black")
    image.save(source)
    package = tmp_path / "sample.document"
    pages = package / "pages"
    pages.mkdir(parents=True)
    source_page = pages / "page-0001.source.png"
    ocr_page = pages / "page-0001.ocr.png"
    source_page.write_bytes(source.read_bytes())
    ocr_page.write_bytes(source.read_bytes())
    words = [
        {"text": "Audit", "box": [100, 100, 130, 45], "confidence": 0.98},
        {"text": "123", "box": [250, 100, 80, 45], "confidence": 0.97},
    ]
    results = [{
        "page": 0,
        "text": "Audit 123",
        "words": words,
        "regions": [
            {"region_type": "heading", "kind": "heading", "text": "Audit 123", "box": [100, 100, 230, 45]},
            {"region_type": "table", "kind": "table", "rows": 2, "cols": 2, "box": [100, 200, 500, 200], "cells": [
                {"text": "Показатель", "row": 0, "col": 0, "rowspan": 1, "colspan": 1},
                {"text": "Значение", "row": 0, "col": 1, "rowspan": 1, "colspan": 1},
                {"text": "Итого", "row": 1, "col": 0, "rowspan": 1, "colspan": 1},
                {"text": "123", "row": 1, "col": 1, "rowspan": 1, "colspan": 1},
            ]},
        ],
        "width": 1200,
        "height": 800,
        "verification": {"status": "pass", "numeric_agreement": 1.0},
        "assets": {
            "source_page_image": "pages/page-0001.source.png",
            "ocr_page_image": "pages/page-0001.ocr.png",
            "source_page_sha256": _sha(source_page),
            "ocr_page_sha256": _sha(ocr_page),
            "raster_dpi": 300,
        },
    }]
    model = create_document_model(
        source,
        results,
        package,
        engine="test-engine",
        engine_params={"api_key": "must-not-leak", "model": "test"},
    )
    assert _sha(package / model.source.path) == _sha(source)
    assert model.metadata["engine_params"]["api_key"] == "[redacted]"
    assert "must-not-leak" not in (package / "document.json").read_text(encoding="utf-8")
    assert load_document_model(package / "document.json").pages[0].primary_text == "Audit 123"

    formats = {"docx", "searchable_pdf", "xlsx", "odt", "markdown", "html", "json", "verification", "archive"}
    files = export_document_model(model, package, tmp_path / "out", formats, stem="sample")
    assert set(files) == formats
    assert "Audit 123" in "\n".join(p.text for p in Document(files["docx"]).paragraphs)
    with fitz.open(files["searchable_pdf"]) as pdf:
        assert len(pdf) == 1
        assert "Audit" in pdf[0].get_text()
        assert pdf[0].get_images()
    workbook = load_workbook(files["xlsx"])
    assert "OCR Audit" in workbook.sheetnames
    assert any(cell.value == "123" for row in workbook["Страница 1"] for cell in row)
    with zipfile.ZipFile(files["odt"]) as archive:
        assert archive.namelist()[0] == "mimetype"
        assert "content.xml" in archive.namelist()
    assert "Audit 123" in files["markdown"].read_text(encoding="utf-8")
    assert "data:image/png;base64," in files["html"].read_text(encoding="utf-8")
    assert json.loads(files["json"].read_text(encoding="utf-8"))["version"] == "1.1"
    with zipfile.ZipFile(files["archive"]) as archive:
        names = archive.namelist()
        assert any(name.startswith("DocumentModel/source/") for name in names)
        assert any(name.startswith("Exports/") for name in names)


def test_column_balancing_prefers_long_text() -> None:
    table = TableRegion(rows=2, cols=3, cells=[
        Cell("№", 0, 0), Cell("Наименование организации и полный адрес объекта", 0, 1), Cell("2025", 0, 2),
        Cell("1", 1, 0), Cell("Очень длинное текстовое значение для проверки ширины", 1, 1), Cell("42", 1, 2),
    ])
    widths = _balanced_column_widths(table, 10.0)
    assert round(sum(widths), 6) == 10.0
    assert widths[1] > widths[0]
    assert widths[1] > widths[2]


def test_mistral_yandex_fusion_requires_native_cell_coordinates(tmp_path: Path) -> None:
    image_path = tmp_path / "grid.png"
    image = Image.new("RGB", (600, 300), "white")
    draw = ImageDraw.Draw(image)
    for x in (50, 300, 550):
        draw.line((x, 50, x, 250), fill="black", width=3)
    for y in (50, 150, 250):
        draw.line((50, y, 550, y), fill="black", width=3)
    image.save(image_path)
    primary = {
        "page": 0,
        "text": "Описание Сумма быстромолодные конструкции 100",
        "mistral_page": {
            "index": 0,
            "dimensions": {"width": 600, "height": 300},
            "blocks": [{
                "type": "table", "table_id": "t1",
                "top_left_x": 50, "top_left_y": 50,
                "bottom_right_x": 550, "bottom_right_y": 250,
            }],
            "tables": [{
                "id": "t1",
                "content": "<table><tr><th>Описание</th><th>Сумма</th></tr><tr><td>быстромолодные конструкции</td><td>100</td></tr></table>",
            }],
        },
        "regions": [{
            "region_type": "table",
            "rows": 2,
            "cols": 2,
            "cells": [
                {"text": "Описание", "row": 0, "col": 0, "box": [50, 50, 250, 100], "coordinate_source": "provider_native"},
                {"text": "Сумма", "row": 0, "col": 1, "box": [300, 50, 250, 100], "coordinate_source": "provider_native"},
                {"text": "быстромолодные конструкции", "row": 1, "col": 0, "box": [50, 150, 250, 100], "coordinate_source": "provider_native"},
                {"text": "100", "row": 1, "col": 1, "box": [300, 150, 250, 100], "coordinate_source": "provider_native"},
            ],
        }],
    }
    yandex = {
        "text": "Описание Сумма быстровозводимые конструкции 101",
        "words": [
            {"text": "Описание", "box": [80, 90, 90, 20], "confidence": 1.0},
            {"text": "Сумма", "box": [360, 90, 70, 20], "confidence": 1.0},
            {"text": "быстровозводимые", "box": [70, 185, 150, 20], "confidence": 1.0},
            {"text": "конструкции", "box": [70, 215, 110, 20], "confidence": 1.0},
            {"text": "101", "box": [370, 195, 40, 20], "confidence": 1.0},
        ],
    }
    fusion = fuse_mistral_yandex(primary, yandex, image_path)
    table = next(region for region in fusion["regions"] if region.get("region_type") == "table")
    values = {(cell["row"], cell["col"]): cell["text"] for cell in table["cells"]}
    assert fusion["applied"]
    assert "быстровозводимые" in values[(1, 0)]
    assert values[(1, 1)] == "100"
    assert fusion["review_items"] == 1

    inferred = copy.deepcopy(primary)
    for cell in inferred["regions"][0]["cells"]:
        cell["coordinate_source"] = "inferred_grid"
    rejected = fuse_mistral_yandex(inferred, yandex, image_path)
    rejected_table = rejected["regions"][0]
    rejected_values = {(cell["row"], cell["col"]): cell["text"] for cell in rejected_table["cells"]}
    assert not rejected["applied"]
    assert rejected_values[(1, 0)] == "быстромолодные конструкции"
    assert rejected["review_items"] == 1
    assert rejected["decisions"][0]["reason"] == "trusted_cell_coordinates_missing"

    tesseract_anchored = copy.deepcopy(primary)
    for cell in tesseract_anchored["regions"][0]["cells"]:
        cell.pop("box", None)
        cell.pop("coordinate_source", None)
    tesseract_anchored["tesseract_2pass"] = {
        "ok": True,
        "words": [
            {"text": "Описание", "box": [80, 90, 90, 20], "confidence": 1.0},
            {"text": "Сумма", "box": [360, 90, 70, 20], "confidence": 1.0},
            {"text": "модулные", "box": [70, 185, 150, 20], "confidence": 1.0},
            {"text": "конструкции", "box": [70, 215, 110, 20], "confidence": 1.0},
            {"text": "100", "box": [370, 195, 40, 20], "confidence": 1.0},
        ],
    }
    tesseract_anchored["regions"][0]["cells"][2]["text"] = "модулные конструкции"
    yandex_anchored = copy.deepcopy(yandex)
    yandex_anchored["words"][2]["text"] = "модульные"
    anchored = fuse_mistral_yandex(tesseract_anchored, yandex_anchored, image_path)
    anchored_table = anchored["regions"][0]
    anchored_values = {(cell["row"], cell["col"]): cell["text"] for cell in anchored_table["cells"]}
    assert anchored["coordinate_anchors"] >= 1
    assert "модульные" in anchored_values[(1, 0)]
    assert anchored_values[(1, 1)] == "100"
    assert needs_yandex_review({"verification": {"status": "pass", "verified": True}}, "suspicious") == (False, "tesseract_pass")


def test_physical_table_geometry_recovers_grid_and_header_merge(tmp_path: Path) -> None:
    image_path = tmp_path / "physical-grid.png"
    image = Image.new("RGB", (620, 280), "white")
    draw = ImageDraw.Draw(image)
    for x in (20, 580):
        draw.line((x, 20, x, 250), fill="black", width=3)
    draw.line((300, 100, 300, 250), fill="black", width=3)
    for y in (20, 100, 130, 250):
        draw.line((20, y, 580, y), fill="black", width=3)
    image.save(image_path)
    regions = [{
        "region_type": "table", "rows": 3, "cols": 2, "box": [20, 20, 560, 230],
        "cells": [
            {"text": "Заголовок", "row": 0, "col": 0, "colspan": 2},
            {"text": "1", "row": 1, "col": 0}, {"text": "2", "row": 1, "col": 1},
            {"text": "А", "row": 2, "col": 0}, {"text": "Б", "row": 2, "col": 1},
        ],
    }]
    words = [
        {"text": "1", "box": [120, 105, 20, 18]},
        {"text": "2", "box": [430, 105, 20, 18]},
    ]
    report = recover_physical_tables(regions, image_path, words)
    assert report[0]["applied"]
    assert report[0]["rows"] == 3 and report[0]["cols"] == 2
    assert report[0]["cells"] == 5
    header = next(cell for cell in regions[0]["cells"] if cell["row"] == 0)
    assert header["colspan"] == 2
    assert header["coordinate_source"] == "physical_grid_verified"


def test_yandex_second_pass_runs_and_can_skip(tmp_path: Path) -> None:
    image_path = tmp_path / "page.png"
    Image.new("RGB", (100, 100), "white").save(image_path)

    class FakeYandex:
        def recognize(self, _path: str, _params: dict) -> OcrResult:
            return OcrResult("yandex", "Документ 123", [Word("Документ", (5, 5, 50, 15), 1.0), Word("123", (60, 5, 20, 15), 1.0)], engine="yandex")

    controller = PipelineController.__new__(PipelineController)
    controller.engines = {}
    controller.root = tmp_path
    controller.tesseract_exe = "tesseract"
    from system_core.ocr_brick.engines import registry

    original = registry.build_adapter
    registry.build_adapter = lambda *_args, **_kwargs: FakeYandex()
    try:
        reviewed = controller._yandex_2pass(
            str(image_path),
            {"yandex_2pass_scope": "all", "yandex_fusion": False},
            {"text": "Документ 124", "verification": {"status": "review", "verified": False}},
        )
        assert reviewed["yandex_2pass"]["ok"]
        assert reviewed["yandex_verification"]["secondary_engine"] == "yandex"
        skipped = controller._yandex_2pass(
            str(image_path),
            {"yandex_2pass_scope": "suspicious", "yandex_fusion": False},
            {"text": "Документ 123", "verification": {"status": "pass", "verified": True}},
        )
        assert skipped["yandex_2pass"]["skipped"]
    finally:
        registry.build_adapter = original
