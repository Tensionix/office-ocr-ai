# render_xlsx.py
# Render brick: OcrResult word boxes -> XLSX grid (one sheet per page).
# Uses openpyxl. Reconstructs rows (line clustering) x columns (x-gap clustering).
# HONEST LIMIT: this is a coordinate heuristic; real tables want a layout engine
# feeding explicit cell regions. Use this for clearly columnar pages. TODO(codex).
# No globals; writes the given path. EN comments only. UTF-8 without BOM.
# Requires: openpyxl.

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

import layout


def render_xlsx(out_path: str | Path,
                pages: list[list[dict]],
                font_name: str = "Arial") -> Path:
    """pages = [ [word_dict, ...], ... ] one word list per page -> one sheet each.
    word_dict = {"text": str, "box": [x,y,w,h], "confidence": float}."""
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet; add per page
    base_font = Font(name=font_name)

    for pi, words in enumerate(pages):
        ws = wb.create_sheet(title=f"page_{pi + 1}")
        grid = layout.build_grid(words)
        for r, row in enumerate(grid, start=1):
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = base_font
        # leave columns at default width; the GUI may offer autosize later

    if not wb.sheetnames:           # no pages -> keep a valid empty workbook
        wb.create_sheet(title="page_1")

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return out
