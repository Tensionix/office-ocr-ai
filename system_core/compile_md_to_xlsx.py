from __future__ import annotations

import os
import re
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from project_paths import ensure_project_dirs, iter_project_files, normalized_source_relative_path, output_path_for

FONT_MAIN = Font(name="Arial", size=10)
FONT_HEADER = Font(name="Arial", size=10, bold=True)
ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
ALIGN_HEADER = Alignment(wrap_text=True, vertical="center", horizontal="center")
FILL_HEADER = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def clean_markdown_cell(text: str) -> str:
    text = re.sub(r"\*\*(.*?)\*\*", r"\1", text.strip())
    text = re.sub(r"\*(.*?)\*", r"\1", text)
    text = re.sub(r"`(.*?)`", r"\1", text)
    return text


def apply_excel_styles(ws) -> None:
    if ws.max_row <= 1 and ws.max_column <= 1 and ws.cell(1, 1).value is None:
        return

    ws.freeze_panes = "A2"
    col_widths = {}

    for row in ws.iter_rows():
        for cell in row:
            if not cell.value:
                continue
            cell.border = THIN_BORDER
            if cell.row == 1:
                cell.font = FONT_HEADER
                cell.alignment = ALIGN_HEADER
                cell.fill = FILL_HEADER
            else:
                cell.font = FONT_MAIN
                cell.alignment = ALIGN_WRAP
            col_widths[cell.column] = max(col_widths.get(cell.column, 0), min(len(str(cell.value)), 60))

    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2


def compile_markdown_to_xlsx() -> None:
    ensure_project_dirs()
    md_files = iter_project_files({".md"})
    if not md_files:
        print("[INFO] No .md files found in project root.")
        return

    print("=== AUDION OFFICE OCR AI: XLSX COMPILER ===")

    for md_file in md_files:
        tables_found = []
        current_table = []
        in_table = False

        for raw_line in md_file.read_text(encoding="utf-8").splitlines():
            line = raw_line.strip()
            if line.startswith("|"):
                if "---" not in line:
                    row_data = [clean_markdown_cell(cell) for cell in line.split("|")[1:-1]]
                    if row_data:
                        current_table.append(row_data)
                        in_table = True
                continue
            if in_table:
                tables_found.append(current_table)
                current_table = []
                in_table = False

        if in_table:
            tables_found.append(current_table)

        if not tables_found:
            continue

        rel_path = normalized_source_relative_path(md_file)
        out_file = output_path_for(md_file, ".xlsx")
        out_file.parent.mkdir(parents=True, exist_ok=True)
        print(f"[BUILD] {rel_path} -> XLSX ({len(tables_found)} table(s))")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for idx, table_data in enumerate(tables_found, start=1):
            ws = wb.create_sheet(title=f"Table_{idx}")
            for row_data in table_data:
                ws.append(row_data)
            apply_excel_styles(ws)

        wb.save(out_file)

    print("\n[OK] Compilation complete.")


if __name__ == "__main__":
    os.system("cls" if os.name == "nt" else "clear")
    compile_markdown_to_xlsx()
